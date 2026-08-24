from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from .paths import FIGURE_SOURCE_ROOT, PUBLICATION_DATA_ROOT, REPOSITORY_ROOT


WORKBOOK = PUBLICATION_DATA_ROOT / "source_data.xlsx"
SOURCES = {
    "Fig02_data": FIGURE_SOURCE_ROOT
    / "main/figure2_relationships/Fig02_relationships_v07.source.csv",
    "Fig03_data": FIGURE_SOURCE_ROOT
    / "main/figure3_empirical_variation/Fig03_empirical_complexity_v05.source.csv",
    "Fig04_data": FIGURE_SOURCE_ROOT
    / "main/figure4_conditional_interventions/Fig04_conditional_interventions_v10.source.csv",
    "Fig05_data": FIGURE_SOURCE_ROOT
    / "main/figure5_open_model_analysis/Fig05_mechanism_evidence_v13.source.csv",
    **{
        f"FigS{number:02d}_data": FIGURE_SOURCE_ROOT
        / f"supplementary/FigS{number:02d}/FigS{number:02d}.source.csv"
        for number in range(1, 13)
    },
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def csv_matrix(path: Path) -> list[list[str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle))


def worksheet_matrix(worksheet) -> list[list[str]]:
    return [
        ["" if cell.value is None else str(cell.value) for cell in row]
        for row in worksheet.iter_rows()
    ]


def validate_source_data(output_dir: Path | None = None) -> dict:
    """Validate all workbook figure sheets against their frozen CSV aliases.

    Equality is checked as a rectangular cell matrix. XLSX and CSV byte hashes
    are intentionally not compared because their container serialization differs.
    """

    workbook = load_workbook(WORKBOOK, read_only=True, data_only=False)
    checks: list[dict] = []
    if output_dir is not None:
        output_dir.mkdir(parents=True, exist_ok=True)
    try:
        for sheet, source in SOURCES.items():
            observed = worksheet_matrix(workbook[sheet])
            expected = csv_matrix(source)
            formulas = sum(
                cell.data_type == "f"
                for row in workbook[sheet].iter_rows()
                for cell in row
            )
            extracted_path = None
            if output_dir is not None:
                extracted_path = output_dir / f"{sheet}.csv"
                with extracted_path.open("w", encoding="utf-8", newline="") as handle:
                    csv.writer(handle, lineterminator="\n").writerows(observed)
            if extracted_path is None:
                extracted_display = None
            else:
                try:
                    extracted_display = extracted_path.resolve().relative_to(
                        REPOSITORY_ROOT.resolve()
                    ).as_posix()
                except ValueError:
                    extracted_display = str(extracted_path.resolve())
            checks.append(
                {
                    "sheet": sheet,
                    "source": source.relative_to(REPOSITORY_ROOT).as_posix(),
                    "source_sha256": sha256(source),
                    "rows_including_header": len(expected),
                    "columns": len(expected[0]) if expected else 0,
                    "cell_matrix_equal": observed == expected,
                    "formula_cells": formulas,
                    "extracted": extracted_display,
                }
            )
    finally:
        workbook.close()
    status = "PASS" if all(item["cell_matrix_equal"] for item in checks) else "FAIL"
    return {
        "schema_version": "1.0",
        "status": status,
        "workbook": WORKBOOK.relative_to(REPOSITORY_ROOT).as_posix(),
        "workbook_sha256": sha256(WORKBOOK),
        "sheet_count": len(checks),
        "cell_matrix_equal_count": sum(item["cell_matrix_equal"] for item in checks),
        "formula_cell_count": sum(item["formula_cells"] for item in checks),
        "checks": checks,
    }


def write_validation_report(report: dict, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
